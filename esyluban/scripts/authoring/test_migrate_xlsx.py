"""Tests for migrate_xlsx.py's row insertion.

The bug these tests exist for: openpyxl's insert_rows moves cell values but
leaves merge ranges where they are. Any cell that lands inside a now-stale
merge range becomes a MergedCell placeholder and its value is discarded --
silently, and specifically for the field-name / type-name header cells.

Five sheets in examples/ carried that damage for months without anyone
noticing, because they were all marked ##export=false and therefore invisible
to the regression baselines.

Run: python test_migrate_xlsx.py
"""
import sys
import tempfile
from pathlib import Path

import openpyxl

# Test names below are Chinese, and Windows hands Python whatever the console
# code page happens to be -- cp936 here, cp1252 on GitHub's runner, where every
# one of these prints raised UnicodeEncodeError and the suite died before its
# first assertion. Pin the stream instead of relying on the environment: it has
# to hold for a contributor's console too, not just for CI.
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).parent))
from migrate_xlsx import ensure_export_row, assert_no_swallowed_values  # noqa: E402

FAILURES = []


def check(name, cond, detail=""):
    if cond:
        print(f"  [OK]   {name}")
    else:
        print(f"  [FAIL] {name}  {detail}")
        FAILURES.append(name)


def build_sheet(tmp, merges, cells):
    wb = openpyxl.Workbook()
    ws = wb.active
    for (r, c), v in cells.items():
        ws.cell(r, c).value = v
    for m in merges:
        ws.merge_cells(start_row=m[0], start_column=m[1], end_row=m[2], end_column=m[3])
    p = Path(tmp) / "t.xlsx"
    wb.save(p)
    return p


def reload(p):
    return openpyxl.load_workbook(p).active


with tempfile.TemporaryDirectory() as tmp:
    # --- 1. horizontal merge on the title row -------------------------------
    p = build_sheet(tmp, [(1, 2, 1, 4)],
                    {(1, 1): "##var", (1, 2): "group", (2, 1): "##type",
                     (2, 2): "int", (3, 1): "", (3, 2): 10})
    ws = reload(p)
    ensure_export_row(ws, 'full_name="a.TbB"', "##export")
    ws.parent.save(p)
    ws = reload(p)
    check("横向合并：值下移一行后仍在",
          ws.cell(2, 2).value == "group", f"实得 {ws.cell(2,2).value!r}")
    check("横向合并：区间跟着下移",
          any((m.min_row, m.min_col, m.max_row, m.max_col) == (2, 2, 2, 4)
              for m in ws.merged_cells.ranges),
          f"实得 {[str(m) for m in ws.merged_cells.ranges]}")
    check("横向合并：##export 写入 A1", ws.cell(1, 1).value == "##export")

    # --- 2. vertical merge, the shape that lost data ------------------------
    # A2:A4 merged with the field name in A2. Before the fix, inserting a row
    # moved 'nums' into A3 which sat inside the stale A2:A4 range -> discarded.
    p = build_sheet(tmp, [(2, 1, 4, 1)],
                    {(1, 1): "##var#column", (2, 1): "nums", (2, 2): "list,int",
                     (2, 4): 11, (3, 4): 12, (4, 4): 13})
    ws = reload(p)
    ensure_export_row(ws, 'full_name="a.TbC"', "##export")
    ws.parent.save(p)
    ws = reload(p)
    check("纵向合并：字段名未被吞掉",
          ws.cell(3, 1).value == "nums", f"实得 {ws.cell(3,1).value!r}")
    check("纵向合并：区间跟着下移",
          any((m.min_row, m.min_col, m.max_row, m.max_col) == (3, 1, 5, 1)
              for m in ws.merged_cells.ranges),
          f"实得 {[str(m) for m in ws.merged_cells.ranges]}")
    check("纵向合并：list 元素完整",
          [ws.cell(r, 4).value for r in (3, 4, 5)] == [11, 12, 13],
          f"实得 {[ws.cell(r,4).value for r in (3,4,5)]}")

    # --- 3. multi-level title, several merges at once -----------------------
    p = build_sheet(tmp, [(1, 2, 1, 3), (1, 4, 1, 5), (2, 2, 2, 3)],
                    {(1, 1): "##var", (1, 2): "x", (1, 4): "y",
                     (2, 1): "##var", (2, 2): "a", (2, 4): "b"})
    ws = reload(p)
    ensure_export_row(ws, None, "##export")
    ws.parent.save(p)
    ws = reload(p)
    got = sorted((m.min_row, m.min_col, m.max_row, m.max_col)
                 for m in ws.merged_cells.ranges)
    check("多级标题：三个合并区全部下移",
          got == [(2, 2, 2, 3), (2, 4, 2, 5), (3, 2, 3, 3)], f"实得 {got}")
    check("多级标题：各级字段名都在",
          (ws.cell(2, 2).value, ws.cell(2, 4).value, ws.cell(3, 2).value) == ("x", "y", "a"))

    # --- 4. idempotence: a sheet already migrated must not shift again ------
    p = build_sheet(tmp, [(2, 2, 2, 4)],
                    {(1, 1): "##export", (1, 2): "full_name=\"a.TbD\"",
                     (2, 1): "##var", (2, 2): "group"})
    ws = reload(p)
    ensure_export_row(ws, 'full_name="a.TbD2"', "##export")
    ws.parent.save(p)
    ws = reload(p)
    check("幂等：已迁移的表不再下移",
          ws.cell(2, 2).value == "group" and ws.cell(1, 2).value == 'full_name="a.TbD2"')

    # --- 5. the guard itself must fire on damage ---------------------------
    wb = openpyxl.Workbook()
    ws2 = wb.active
    ws2.cell(2, 1).value = "kept"
    ws2.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)  # empty top-left
    fired = False
    try:
        assert_no_swallowed_values(ws2, "test")
    except RuntimeError:
        fired = True
    check("自检闸门：左上角为空时报错", fired)

    # ── main() 的元数据逻辑 ────────────────────────────────────────
    #
    # 上面十项测的都是 ensure_export_row，也就是「怎么插这一行」。而这个工具
    # 被禁用的理由是另一半：main() 里从 __tables__.xlsx 读出元数据、拼成 B1
    # 的那段，因为一个缩进错误长期不可达，从未在真实环境跑过。
    #
    # 这里造一份最小的上游格式工程跑完整流程，覆盖那段逻辑的四条分支：
    # 登记在 __tables__ 里的表、__beans__ 定义文件、__tables__ 自身、
    # 以及带合并单元格的表（历史上「插行吞数据」的触发点）。
    proj = Path(tmp) / "proj"
    (proj / "DataTables" / "item").mkdir(parents=True)

    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["##var", "full_name", "value_type", "index", "mode", "group", "comment", "input"])
    ws.append(["##type", "string", "string", "string", "string", "string", "string", "string"])
    ws.append(["##", "", "", "", "", "", "", ""])
    ws.append(["", "item.TbItem", "item.Item", "id", "", "", "道具表", "item/items.xlsx"])
    wb.save(proj / "DataTables" / "__tables__.xlsx")

    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["##var", "id", "name"])
    ws.append(["##type", "int", "string"])
    ws.append(["", 1, "Sword"])
    ws.merge_cells("B1:C1")
    wb.save(proj / "DataTables" / "item" / "items.xlsx")

    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["##var", "full_name"])
    ws.append(["", "item.Item"])
    wb.save(proj / "DataTables" / "__beans__.xlsx")

    (proj / "Tools" / "Luban").mkdir(parents=True)
    (proj / "Tools" / "Luban" / "luban.conf").write_text(
        '{"schemaFiles":[{"fileName":"../../DataTables/__tables__.xlsx","type":"table"}],'
        '"dataDir":"../../DataTables","groups":[],"targets":[],"xargs":[]}',
        encoding="utf-8")

    import os
    import subprocess
    env = dict(os.environ, LUBAN_ROOT=str(proj))
    src = (Path(__file__).parent / "migrate_xlsx.py").read_text(encoding="utf-8")
    gate_start = src.index('if __name__ == "__main__":')
    gate_end = src.index("raise SystemExit(2)") + len("raise SystemExit(2)")
    runner = Path(tmp) / "run_migrate.py"
    runner.write_text(src[:gate_start] + src[gate_end + 1:], encoding="utf-8")
    # encoding 要显式给：text=True 默认按 locale 解码子进程输出，而这个子进程
    # 打印中文。r.stderr 会被塞进下面那条断言的失败信息里 —— 解码炸掉的话，
    # 报错会指向解码而不是指向真正失败的那一项。
    r = subprocess.run([sys.executable, str(runner)], env=env, cwd=str(proj),
                       capture_output=True, text=True,
                       encoding="utf-8", errors="replace")

    check("main：完整迁移跑通", r.returncode == 0, r.stderr[-200:])

    items = openpyxl.load_workbook(proj / "DataTables" / "item" / "items.xlsx").active
    check("main：元数据从 __tables__ 拼进 B1",
          isinstance(items.cell(1, 2).value, str)
          and 'full_name="item.TbItem"' in items.cell(1, 2).value
          and 'comment="道具表"' in items.cell(1, 2).value,
          repr(items.cell(1, 2).value))
    check("main：合并区跟着下移", "B2:C2" in [str(m) for m in items.merged_cells.ranges],
          str([str(m) for m in items.merged_cells.ranges]))
    check("main：数据行未被吞掉",
          items.cell(4, 2).value == 1 and items.cell(4, 3).value == "Sword",
          f"{items.cell(4, 2).value!r} {items.cell(4, 3).value!r}")

    beans = openpyxl.load_workbook(proj / "DataTables" / "__beans__.xlsx").active
    check("main：__beans__ 被标记为导出", beans.cell(1, 1).value == "##export",
          repr(beans.cell(1, 1).value))

    tables = openpyxl.load_workbook(proj / "DataTables" / "__tables__.xlsx").active
    check("main：__tables__ 自身标记为不导出",
          tables.cell(1, 1).value == "##export=false", repr(tables.cell(1, 1).value))

print()
if FAILURES:
    print(f"FAILED: {len(FAILURES)} 项 -> {FAILURES}")
    sys.exit(1)
print("ALL PASSED")
