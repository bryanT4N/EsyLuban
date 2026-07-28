"""一次性迁移工具：把上游集中式 __tables__.xlsx 的表定义改写为自包含格式。

⚠ 仍处于禁用状态，但原因已经变了 —— 请读完再决定是否解禁。

【已修复，且有测试保证】

  1. ensure_export_row() 插行时不再吞数据。
     openpyxl 的 insert_rows 只移动值、不移动合并区间，落进陈旧合并区间的
     单元格会被写成 MergedCell 占位、**值被永久删除**，丢的恰好是字段名与
     类型名这类标题格。现在改为「先快照并全部 unmerge → insert_rows →
     按新坐标重新 merge」，并在保存前自检每个合并区左上角非空。
     这个 bug 曾在 examples 里造成 5 处损伤（20 个单元格），已按上游
     luban_examples@879f5c5 逐单元格审计并补回。
     回归测试见同目录 test_migrate_xlsx.py（退回旧实现会有 4 项失败）。

  2. main() 的循环缩进已修正。
     原本 `for ws in wb.worksheets` 与 `for name in filenames` 平级，
     每个目录只跑一次且复用上一次的 wb，首个目录没有 xlsx 便直接 NameError。

  3. main() 的元数据逻辑已跑通并有测试。
     349 行起那段（meta_map / build_autoimport_meta / xml_map / export_false）
     曾因缩进错误从未执行过。现在 test_migrate_xlsx.py 用一份最小的上游格式
     工程跑完整流程，覆盖四条分支：登记在 __tables__ 里的表、__beans__ 定义
     文件、__tables__ 自身、以及带合并单元格的表。
     另在一份三表语料上实跑验证过：60 个非空单元格零丢失，合并区 B1:C1 正确
     下移到 B2:C2，元数据（含 comment 与 input）完整拼进 B1。

【为什么仍然禁用】

  上面这些证明的是「它在我构造的语料上是对的」，不是「它在你的工程上是对的」。
  真实上游工程会有多 sheet、CSV、XML 定义、跨目录 input、多套 luban.conf ——
  这些组合没有一个被验证过。

  而它失败的代价不可逆：它【原地改写】你的 Excel，没有备份、没有 dry-run。
  一个只在部分情况下正确的原地改写工具，比没有工具更危险。

要用它：删除下方的 SystemExit，**先把整个工程复制一份**，在副本上跑，与原件
逐单元格比对确认无损，再决定是否用于正式工程。当前 examples 已是迁移完成
状态，不需要再跑它。
"""
import csv
import json
import os
import pathlib
import sys
from datetime import datetime
import xml.etree.ElementTree as ET

from openpyxl import load_workbook

# 必须在下面那道闸门之前：闸门的警告文本是中文，而 Windows 给 Python 的编码是
# 当前控制台代码页（本机 cp936，GitHub runner 是 cp1252）。编不出来时抛的是
# UnicodeEncodeError —— 一道本该说清「为什么不让你跑」的闸门，反而以看不懂的
# 报错收场。
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

# 拒绝执行优于文档警告：注释救不了直接双击运行的人。
if __name__ == "__main__":
    sys.stderr.write(
        "\n[DISABLED] migrate_xlsx.py 不可直接运行。\n"
        "  它【原地改写】你的 Excel —— 没有备份，也没有 dry-run。\n"
        "  已修的缺陷与已覆盖的测试见 test_migrate_xlsx.py（16 项）。\n"
        "  但多 sheet / CSV / XML 定义 / 多套 conf 这些组合从未验证过。\n"
        "  要用它：先复制整个工程，在副本上跑，逐单元格比对确认无损。\n"
        "  详见本文件顶部说明。\n\n")
    raise SystemExit(2)


ROOT = os.environ.get("LUBAN_ROOT", r"C:\Users\BryanT\Documents\WORK_PROJECTS\APP_PROJECTS\EsyLuban")
ARCHIVE_DIR = os.path.join(ROOT, "TestOutputs", "migration_20260122")
os.makedirs(ARCHIVE_DIR, exist_ok=True)
REPORT_PATH = os.path.join(ARCHIVE_DIR, "migration_report.json")

TABLE_FIELDS_ORDER = [
    "full_name",
    "value_type",
    "index",
    "mode",
    "group",
    "comment",
    "read_schema_from_file",
    "input",
    "output",
    "tags",
]


def split_file_and_sheet(url: str):
    if "@" not in url:
        return url, None
    sheet_sep = url.find("@")
    last_path_sep = url.rfind("/", 0, sheet_sep)
    if last_path_sep >= 0:
        file_path = url[: last_path_sep + 1] + url[sheet_sep + 1 :]
        sheet_name = url[last_path_sep + 1 : sheet_sep]
        return file_path, sheet_name
    return url[sheet_sep + 1 :], url[:sheet_sep]


def norm_path(p: str):
    return os.path.normpath(p).replace("\\", "/")


def compose_meta(record: dict):
    pairs = []
    for key in TABLE_FIELDS_ORDER:
        value = record.get(key, "")
        if value is None:
            value = ""
        value = str(value).strip()
        if not value:
            continue
        value = value.replace("\\", "\\\\").replace("\"", "\\\"")
        pairs.append(f'{key}="{value}"')
    return " & ".join(pairs)


def build_autoimport_meta(file_path: str, data_dir: str):
    if not data_dir:
        return None
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    if not base_name.startswith("#"):
        return None
    raw = base_name[1:]
    if "-" in raw:
        raw = raw.split("-", 1)[0]
    if not raw:
        return None
    if "." in raw:
        raw_namespace, raw_name = raw.rsplit(".", 1)
    else:
        raw_namespace, raw_name = "", raw
    rel = os.path.relpath(file_path, data_dir).replace("\\", "/")
    ns_from_path = os.path.dirname(rel).replace("/", ".").replace("\\", ".")
    ns_from_path = "" if ns_from_path == "." else ns_from_path
    ns = ".".join([p for p in [ns_from_path, raw_namespace] if p])
    table_name = f"Tb{raw_name}"
    if ns:
        full_name = f"{ns}.{table_name}"
        value_type = f"{ns}.{raw_name}"
    else:
        full_name = table_name
        value_type = raw_name
    record = {
        "full_name": full_name,
        "value_type": value_type,
        "index": "",
        "mode": "map",
        "group": "",
        "comment": "",
        "read_schema_from_file": "true",
        "input": rel,
        "output": "",
        "tags": "",
    }
    return compose_meta(record)


def is_excel_file(path: str):
    lower = path.lower()
    return lower.endswith(".xlsx") or lower.endswith(".xls") or lower.endswith(".xlsm") or lower.endswith(".csv")


def parse_tables_xlsx(tables_path: str):
    wb = load_workbook(tables_path, data_only=True)
    table_defs = []
    for ws in wb.worksheets:
        header_row_idx = None
        for r in range(1, ws.max_row + 1):
            v = ws.cell(r, 1).value
            if isinstance(v, str) and v.strip().lower() == "##var":
                header_row_idx = r
                break
        if header_row_idx is None:
            continue
        headers = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row_idx, c).value
            headers.append((v or "").strip() if isinstance(v, str) else (v or ""))
        data_start = header_row_idx + 1
        while data_start <= ws.max_row:
            first = ws.cell(data_start, 1).value
            if isinstance(first, str) and first.strip().startswith("##"):
                data_start += 1
                continue
            break
        for r in range(data_start, ws.max_row + 1):
            first = ws.cell(r, 1).value
            if isinstance(first, str) and first.strip().startswith("##"):
                break
            row_vals = {}
            empty = True
            for c, h in enumerate(headers, start=1):
                if not h:
                    continue
                v = ws.cell(r, c).value
                if v is None:
                    v = ""
                if isinstance(v, str):
                    v = v.strip()
                row_vals[h] = v
                if str(v).strip():
                    empty = False
            if empty:
                continue
            table_defs.append(row_vals)
    return table_defs


def build_meta_map():
    table_meta = {}
    table_meta_anysheet = {}
    unresolved = []
    tables_files = []
    for dirpath, _, filenames in os.walk(ROOT):
        if "bin" in pathlib.Path(dirpath).parts or "obj" in pathlib.Path(dirpath).parts:
            continue
        for name in filenames:
            if name.lower() == "__tables__.xlsx":
                tables_files.append(os.path.join(dirpath, name))
    tables_files.sort()

    for tables_path in tables_files:
        data_dir = os.path.dirname(tables_path)
        for record in parse_tables_xlsx(tables_path):
            input_raw = str(record.get("input", "") or "").strip()
            full_name = str(record.get("full_name", "") or "").strip()
            if not full_name:
                continue
            meta_str = compose_meta(record)
            inputs = [s.strip() for s in input_raw.split(",") if s.strip()]
            excel_inputs = []
            for inp in inputs:
                file_part, sheet_part = split_file_and_sheet(norm_path(inp))
                abs_path = os.path.join(data_dir, file_part)
                if is_excel_file(file_part):
                    excel_inputs.append((os.path.normpath(abs_path), sheet_part))
            if not excel_inputs:
                unresolved.append({"table": full_name, "input": input_raw, "source": tables_path})
                continue
            target = excel_inputs[0]
            if target[1] is None:
                table_meta_anysheet.setdefault(target[0], []).append({"meta": meta_str, "table": full_name, "source": tables_path})
            else:
                table_meta.setdefault(target, []).append({"meta": meta_str, "table": full_name, "source": tables_path})
    return table_meta, table_meta_anysheet, unresolved


def has_header(ws):
    for r in range(1, min(ws.max_row, 6) + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower().startswith("##export"):
            continue
        if isinstance(v, str) and v.strip().startswith("##"):
            return True
    return False


def load_luban_confs():
    confs = []
    for dirpath, _, filenames in os.walk(ROOT):
        if "bin" in pathlib.Path(dirpath).parts or "obj" in pathlib.Path(dirpath).parts:
            continue
        for name in filenames:
            if name.lower() == "luban.conf":
                conf_path = os.path.join(dirpath, name)
                try:
                    with open(conf_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    data_dir = data.get("dataDir", "")
                    if data_dir:
                        abs_data_dir = os.path.normpath(os.path.join(dirpath, data_dir))
                        confs.append({"conf_dir": os.path.normpath(dirpath), "data_dir": abs_data_dir})
                except Exception:
                    continue
    confs.sort(key=lambda x: len(x["conf_dir"]), reverse=True)
    return confs


def find_data_dir(confs, file_path):
    file_path = os.path.normpath(file_path)
    for c in confs:
        if file_path.startswith(c["data_dir"] + os.sep):
            return c["data_dir"]
        if file_path.startswith(c["conf_dir"] + os.sep):
            return c["data_dir"]
    return None


def collect_xml_excel_inputs(confs):
    xml_map = {}
    xml_anysheet = {}
    xml_dirs = set()
    for dirpath, _, filenames in os.walk(ROOT):
        if "bin" in pathlib.Path(dirpath).parts or "obj" in pathlib.Path(dirpath).parts:
            continue
        for name in filenames:
            if not name.lower().endswith(".xml"):
                continue
            full_path = os.path.join(dirpath, name)
            data_dir = find_data_dir(confs, full_path)
            if not data_dir:
                continue
            try:
                tree = ET.parse(full_path)
            except Exception:
                continue
            for elem in tree.iter():
                if elem.tag != "table":
                    continue
                input_raw = (elem.attrib.get("input") or "").strip()
                if not input_raw:
                    continue
                inputs = [s.strip() for s in input_raw.split(",") if s.strip()]
                for inp in inputs:
                    file_part, sheet_part = split_file_and_sheet(norm_path(inp))
                    abs_path = os.path.normpath(os.path.join(data_dir, file_part))
                    if os.path.isdir(abs_path):
                        xml_dirs.add(abs_path)
                        continue
                    if not is_excel_file(file_part):
                        continue
                    if sheet_part:
                        xml_map.setdefault((abs_path, sheet_part), []).append(full_path)
                    else:
                        xml_anysheet.setdefault(abs_path, []).append(full_path)
    return xml_map, xml_anysheet, xml_dirs


def shift_merges_down(ws, by: int = 1):
    """把全部合并区间整体下移 by 行。

    必须在 insert_rows 之前调用，且顺序不能反。openpyxl 的 insert_rows 只移动
    单元格的值，合并区间原地不动；于是下移后落进陈旧合并区间的单元格会被写成
    MergedCell 占位，**值被永久丢弃** —— 丢的恰好是字段名、类型名这类标题格。

    正确顺序：先快照并全部 unmerge，再 insert_rows，最后按新坐标重新 merge。
    """
    snapshot = [(m.min_row, m.min_col, m.max_row, m.max_col)
                for m in ws.merged_cells.ranges]
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))
    return snapshot


def reapply_merges(ws, snapshot, by: int = 1):
    for r1, c1, r2, c2 in snapshot:
        ws.merge_cells(start_row=r1 + by, start_column=c1,
                       end_row=r2 + by, end_column=c2)


def assert_no_swallowed_values(ws, where: str):
    """插行后自检：每个合并区的左上角都必须有值。

    左上角为空意味着这个合并区不再对应任何内容 —— 要么位移错了，要么值被吞了。
    与其把损坏写进文件让人半年后才发现，不如当场失败。
    """
    for m in ws.merged_cells.ranges:
        if ws.cell(m.min_row, m.min_col).value in (None, ""):
            raise RuntimeError(
                f"{where}: 合并区 {m} 的左上角为空，插行后值可能已被丢弃，已中止")


def ensure_export_row(ws, meta: str, export_flag: str):
    a1 = ws.cell(1, 1).value
    if isinstance(a1, str) and a1.strip().lower().startswith("##export"):
        ws.cell(1, 1).value = export_flag
        if meta is not None:
            ws.cell(1, 2).value = meta
        return
    snapshot = shift_merges_down(ws)
    ws.insert_rows(1)
    reapply_merges(ws, snapshot)
    ws.cell(1, 1).value = export_flag
    if meta is not None:
        ws.cell(1, 2).value = meta
    assert_no_swallowed_values(ws, f"sheet:{ws.title}")


def migrate_csv(path: str, meta: str, export_flag: str):
    encodings = ["utf-8-sig", "utf-8", "gbk", "latin1"]
    rows = None
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                rows = list(csv.reader(f))
            break
        except UnicodeDecodeError:
            continue
    if rows is None:
        raise UnicodeDecodeError("utf-8", b"", 0, 1, "unsupported encoding")
    new_header = [export_flag, meta or ""]
    if rows:
        first = rows[0][0].strip().lower() if rows[0] else ""
        if first.startswith("##export"):
            rows[0] = new_header + rows[0][2:]
        else:
            rows.insert(0, new_header)
    else:
        rows = [new_header]
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)


def main():
    meta_map, meta_anysheet, unresolved = build_meta_map()
    confs = load_luban_confs()
    xml_map, xml_anysheet, xml_dirs = collect_xml_excel_inputs(confs)
    migrated = []
    export_false = []
    skipped = []

    for dirpath, _, filenames in os.walk(ROOT):
        if "bin" in pathlib.Path(dirpath).parts or "obj" in pathlib.Path(dirpath).parts:
            continue
        for name in filenames:
            lower = name.lower()
            if not (lower.endswith(".xlsx") or lower.endswith(".xlsm") or lower.endswith(".xls") or lower.endswith(".csv")):
                continue
            full_path = os.path.join(dirpath, name)
            rel_path = os.path.relpath(full_path, ROOT)
            data_dir = find_data_dir(confs, full_path)
            if lower.endswith(".csv"):
                file_key = os.path.normpath(full_path)
                candidates = [(k, v) for k, v in meta_map.items() if os.path.normpath(k[0]) == file_key]
                if candidates:
                    meta = candidates[0][1][0]["meta"]
                    migrate_csv(full_path, meta, "##export")
                    migrated.append({"file": rel_path, "sheet": None, "meta": meta})
                elif file_key in xml_anysheet or (file_key, None) in xml_map or any(file_key.startswith(d + os.sep) for d in xml_dirs):
                    migrate_csv(full_path, "", "##export")
                    migrated.append({"file": rel_path, "sheet": None, "meta": ""})
                else:
                    skipped.append({"file": rel_path, "reason": "no_meta"})
                continue
            wb = load_workbook(full_path)
            is_def_file = lower in ("__beans__.xlsx", "__enums__.xlsx")
            # 这个循环原本缩进错了一级，与 `for name in filenames` 平级，
            # 于是每个目录只跑一次、用的还是上一次 load_workbook 留下的 wb
            # （os.walk 的第一个目录没有 xlsx，因此直接 NameError）。
            for ws in wb.worksheets:
                if ws.title.lower() in ("__beans__", "__enums__"):
                    ensure_export_row(ws, None, "##export")
                    migrated.append({"file": rel_path, "sheet": ws.title, "meta": ""})
                    continue
                key = (os.path.normpath(full_path), ws.title)
                if key in meta_map:
                    meta = meta_map[key][0]["meta"]
                    ensure_export_row(ws, meta, "##export")
                    migrated.append({"file": rel_path, "sheet": ws.title, "meta": meta})
                elif build_autoimport_meta(full_path, data_dir) and has_header(ws):
                    meta = build_autoimport_meta(full_path, data_dir)
                    ensure_export_row(ws, meta, "##export")
                    migrated.append({"file": rel_path, "sheet": ws.title, "meta": meta})
                elif os.path.normpath(full_path) in meta_anysheet and has_header(ws):
                    meta = meta_anysheet[os.path.normpath(full_path)].pop(0)["meta"]
                    ensure_export_row(ws, meta, "##export")
                    migrated.append({"file": rel_path, "sheet": ws.title, "meta": meta})
                    if not meta_anysheet[os.path.normpath(full_path)]:
                        meta_anysheet.pop(os.path.normpath(full_path), None)
                elif key in xml_map and has_header(ws):
                    ensure_export_row(ws, "", "##export")
                    migrated.append({"file": rel_path, "sheet": ws.title, "meta": ""})
                elif os.path.normpath(full_path) in xml_anysheet and has_header(ws):
                    ensure_export_row(ws, "", "##export")
                    migrated.append({"file": rel_path, "sheet": ws.title, "meta": ""})
                elif any(os.path.normpath(full_path).startswith(d + os.sep) for d in xml_dirs) and has_header(ws):
                    ensure_export_row(ws, "", "##export")
                    migrated.append({"file": rel_path, "sheet": ws.title, "meta": ""})
                elif is_def_file:
                    ensure_export_row(ws, None, "##export")
                    migrated.append({"file": rel_path, "sheet": ws.title, "meta": ""})
                else:
                    a1 = ws.cell(1, 1).value
                    if isinstance(a1, str) and a1.strip().startswith("##"):
                        if isinstance(a1, str) and a1.strip().lower().startswith("##export") and not has_header(ws):
                            ensure_export_row(ws, None, "##export=false")
                            export_false.append({"file": rel_path, "sheet": ws.title})
                        else:
                            ensure_export_row(ws, None, "##export=false")
                            export_false.append({"file": rel_path, "sheet": ws.title})
                    else:
                        skipped.append({"file": rel_path, "sheet": ws.title, "reason": "no_header"})
            wb.save(full_path)

    report = {
        "timestamp": datetime.now().isoformat(),
        "migrated": migrated,
        "export_false": export_false,
        "skipped": skipped,
        "unresolved_tables": unresolved,
        "unassigned_anysheet": meta_anysheet,
        "xml_dirs": sorted(xml_dirs),
    }
    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"migrated:{len(migrated)} export_false:{len(export_false)} skipped:{len(skipped)} unresolved:{len(unresolved)}")


if __name__ == "__main__":
    main()
