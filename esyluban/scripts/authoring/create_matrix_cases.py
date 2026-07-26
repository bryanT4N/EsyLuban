import argparse
from pathlib import Path

import openpyxl


def set_table_sheet(
    ws,
    full_name,
    value_type,
    mode="map",
    index="id",
    extra_meta=None,
    read_schema_from_file=True,
):
    ws.cell(1, 1).value = "##export"
    meta = f'full_name="{full_name}" & value_type="{value_type}"'
    if index:
        meta += f' & index="{index}"'
    if mode:
        meta += f' & mode="{mode}"'
    if not read_schema_from_file:
        meta += ' & read_schema_from_file="false"'
    if extra_meta:
        meta += f" & {extra_meta}"
    ws.cell(1, 2).value = meta
    ws.cell(2, 1).value = "##var"
    ws.cell(3, 1).value = "##type"


def write_row(ws, row_index, values, start_col=2):
    for idx, value in enumerate(values, start=start_col):
        ws.cell(row_index, idx).value = value


def build_enums_sheet(ws):
    ws.cell(1, 1).value = "##export"
    ws.cell(2, 1).value = "##var"
    write_row(ws, 2, ["full_name", "flags", "unique", "group", "comment", "tags", "*items"])
    ws.cell(3, 1).value = "##var"
    write_row(ws, 3, ["", "", "", "", "", "", "name", "alias", "value", "comment", "tags"])
    ws.cell(4, 1).value = "##"
    write_row(
        ws,
        5,
        ["matrix.EQuality", False, True, None, "quality enum", None, "Common", "Common", 1, "Common", None],
    )
    write_row(ws, 6, [None, None, None, None, None, None, "Rare", "Rare", 2, "Rare", None])


def build_beans_sheet(ws, beans):
    ws.cell(1, 1).value = "##export"
    ws.cell(2, 1).value = "##var"
    write_row(
        ws,
        2,
        [
            "full_name",
            "parent",
            "valueType",
            "alias",
            "sep",
            "comment",
            "tags",
            "group",
            "*fields",
        ],
    )
    ws.cell(3, 1).value = "##var"
    write_row(ws, 3, ["", "", "", "", "", "", "", "", "name", "alias", "type", "group", "comment", "tags"])
    ws.cell(4, 1).value = "##"

    row = 5
    for bean in beans:
        first = True
        for field in bean["fields"]:
            if first:
                write_row(
                    ws,
                    row,
                    [
                        bean["full_name"],
                        bean.get("parent", ""),
                        bean.get("valueType", False),
                        bean.get("alias", ""),
                        bean.get("sep", ""),
                        bean.get("comment", ""),
                        bean.get("tags", ""),
                        bean.get("group", ""),
                        field["name"],
                        field.get("alias", field["name"]),
                        field["type"],
                        field.get("group", ""),
                        field.get("comment", ""),
                        field.get("tags", ""),
                        "",
                    ],
                )
                first = False
            else:
                write_row(
                    ws,
                    row,
                    [
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        "",
                        field["name"],
                        field.get("alias", field["name"]),
                        field["type"],
                        field.get("group", ""),
                        field.get("comment", ""),
                        field.get("tags", ""),
                        "",
                    ],
                )
            row += 1


def build_feature_tables(data_root: Path):
    matrix_dir = data_root / "matrix"
    matrix_dir.mkdir(parents=True, exist_ok=True)
    path = matrix_dir / "feature_tables.xlsx"

    wb = openpyxl.Workbook()
    ws_basic = wb.active
    ws_basic.title = "basic"
    set_table_sheet(ws_basic, "matrix.TbBasic", "matrix.BasicRecord", mode="map", index="id", read_schema_from_file=False)
    write_row(
        ws_basic,
        2,
        [
            "id",
            "name",
            "quality",
            "textKey",
            "res",
            "enabled",
            "score",
            "count",
            "created",
            "optName",
            "notDefault",
        ],
    )
    write_row(
        ws_basic,
        3,
        [
            "int",
            "string",
            "matrix.EQuality",
            "text",
            "string#(path=unity)",
            "bool",
            "float",
            "long",
            "datetime",
            "string?",
            "int!",
        ],
    )
    write_row(
        ws_basic,
        4,
        [
            1,
            "Alpha",
            "Common",
            "/apple",
            "Scenes/SampleScene.unity",
            True,
            1.5,
            100,
            "2025-01-01 00:00:00",
            None,
            1,
        ],
    )

    ws_containers = wb.create_sheet("containers")
    set_table_sheet(ws_containers, "matrix.TbContainers", "matrix.ContainerRecord", mode="map", index="id", read_schema_from_file=False)
    write_row(ws_containers, 2, ["id", "nums#sep=,", "tags#sep=,", "props#sep=,", "innerList#sep=,"])
    write_row(ws_containers, 3, ["int", "list,int", "set,string", "map,int,string", "list,matrix.InnerBean"])
    write_row(ws_containers, 4, [1, "1,2,3", "A,B", "1,apple,2,banana", "1:Alpha,2:Beta"])

    ws_singleton = wb.create_sheet("singleton")
    set_table_sheet(ws_singleton, "matrix.TbMatrixSingleton", "matrix.SingleConfig", mode="one", index=None, read_schema_from_file=False)
    write_row(ws_singleton, 2, ["version", "title"])
    write_row(ws_singleton, 3, ["int", "string"])
    write_row(ws_singleton, 4, [1, "Config"])

    ws_list = wb.create_sheet("list")
    set_table_sheet(ws_list, "matrix.TbMatrixList", "matrix.ListRecord", mode="list", index=None, read_schema_from_file=False)
    write_row(ws_list, 2, ["id", "name"])
    write_row(ws_list, 3, ["int", "string"])
    write_row(ws_list, 4, [1, "A"])
    write_row(ws_list, 5, [2, "B"])

    wb.save(path)
    return path


def build_negative_path_case(data_root: Path):
    neg_dir = data_root / "negatives"
    neg_dir.mkdir(parents=True, exist_ok=True)
    path = neg_dir / "path_fail.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    set_table_sheet(
        ws,
        "matrix.TbPathFail",
        "matrix.PathRecord",
        mode="map",
        index="id",
        extra_meta='group="t"',
        read_schema_from_file=False,
    )
    write_row(ws, 2, ["id", "res"])
    write_row(ws, 3, ["int", "string#(path=unity)"])
    write_row(ws, 4, [1, "Scenes/NotExist.unity"])

    wb.save(path)
    return path


def append_enum_defs(data_root: Path):
    enum_path = data_root / "__enums__.xlsx"
    wb = openpyxl.load_workbook(enum_path)
    ws = wb.active

    existing = set()
    for r in range(5, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if isinstance(val, str) and val.strip():
            existing.add(val.strip())

    if "matrix.EQuality" in existing:
        for r in range(5, ws.max_row + 1):
            if ws.cell(r, 2).value == "matrix.EQuality":
                ws.cell(r, 9).value = None
                ws.cell(r + 1, 9).value = None
                break
        wb.save(enum_path)
        return

    row = ws.max_row + 1
    ws.cell(row, 2).value = "matrix.EQuality"
    ws.cell(row, 3).value = False
    ws.cell(row, 4).value = True
    ws.cell(row, 6).value = "quality enum"
    ws.cell(row, 8).value = "Common"
    ws.cell(row, 9).value = None
    ws.cell(row, 10).value = 1
    ws.cell(row, 11).value = "Common"
    row += 1
    ws.cell(row, 8).value = "Rare"
    ws.cell(row, 9).value = None
    ws.cell(row, 10).value = 2
    ws.cell(row, 11).value = "Rare"

    wb.save(enum_path)


def append_bean_defs(data_root: Path):
    bean_path = data_root / "__beans__.xlsx"
    wb = openpyxl.load_workbook(bean_path)
    ws = wb.active

    existing = set()
    for r in range(5, ws.max_row + 1):
        val = ws.cell(r, 2).value
        if isinstance(val, str) and val.strip():
            existing.add(val.strip())

    beans = [
        {
            "full_name": "matrix.BasicRecord",
            "comment": "basic record",
            "fields": [
                {"name": "id", "type": "int"},
                {"name": "name", "type": "string"},
                {"name": "quality", "type": "matrix.EQuality"},
                {"name": "textKey", "type": "text"},
                {"name": "res", "type": "string#(path=unity)"},
                {"name": "enabled", "type": "bool"},
                {"name": "score", "type": "float"},
                {"name": "count", "type": "long"},
                {"name": "created", "type": "datetime"},
                {"name": "optName", "type": "string?"},
                {"name": "notDefault", "type": "int!"},
            ],
        },
        {
            "full_name": "matrix.InnerBean",
            "sep": ":",
            "comment": "inner bean",
            "fields": [
                {"name": "id", "type": "int"},
                {"name": "name", "type": "string"},
            ],
        },
        {
            "full_name": "matrix.ContainerRecord",
            "comment": "container record",
            "fields": [
                {"name": "id", "type": "int"},
                {"name": "nums", "type": "list,int"},
                {"name": "tags", "type": "set,string"},
                {"name": "props", "type": "map,int,string"},
                {"name": "innerList", "type": "list,matrix.InnerBean"},
            ],
        },
        {
            "full_name": "matrix.SingleConfig",
            "comment": "singleton config",
            "fields": [
                {"name": "version", "type": "int"},
                {"name": "title", "type": "string"},
            ],
        },
        {
            "full_name": "matrix.ListRecord",
            "comment": "list record",
            "fields": [
                {"name": "id", "type": "int"},
                {"name": "name", "type": "string"},
            ],
        },
        {
            "full_name": "matrix.PathRecord",
            "comment": "path record",
            "fields": [
                {"name": "id", "type": "int"},
                {"name": "res", "type": "string#(path=unity)"},
            ],
        },
    ]

    row = ws.max_row + 1
    for bean in beans:
        if bean["full_name"] in existing:
            for r in range(5, ws.max_row + 1):
                if ws.cell(r, 2).value == bean["full_name"]:
                    rr = r
                    while rr <= ws.max_row:
                        if rr != r and ws.cell(rr, 2).value:
                            break
                        if ws.cell(rr, 10).value:
                            ws.cell(rr, 11).value = None
                        rr += 1
                    break
            continue
        first = True
        for field in bean["fields"]:
            if first:
                ws.cell(row, 2).value = bean["full_name"]
                ws.cell(row, 6).value = bean.get("sep", "")
                ws.cell(row, 7).value = bean.get("comment", "")
                ws.cell(row, 10).value = field["name"]
                ws.cell(row, 11).value = field.get("alias", "")
                ws.cell(row, 12).value = field["type"]
                ws.cell(row, 14).value = field.get("comment", "")
                first = False
            else:
                ws.cell(row, 10).value = field["name"]
                ws.cell(row, 11).value = field.get("alias", "")
                ws.cell(row, 12).value = field["type"]
                ws.cell(row, 14).value = field.get("comment", "")
            row += 1

    wb.save(bean_path)


def ensure_assets(data_root: Path):
    asset_file = data_root / "Assets" / "Scenes" / "SampleScene.unity"
    asset_file.parent.mkdir(parents=True, exist_ok=True)
    asset_file.write_text("", encoding="utf-8")
    return asset_file


def main() -> int:
    parser = argparse.ArgumentParser(description="Create feature matrix tables and negative cases.")
    parser.add_argument("--data-root", required=True, help="DataTables root directory")
    args = parser.parse_args()

    data_root = Path(args.data_root)
    append_enum_defs(data_root)
    append_bean_defs(data_root)
    build_feature_tables(data_root)
    build_negative_path_case(data_root)
    ensure_assets(data_root)
    print("matrix cases updated")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
