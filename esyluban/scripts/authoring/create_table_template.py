import argparse
from pathlib import Path

import openpyxl


def main() -> int:
    parser = argparse.ArgumentParser(description="Create a minimal EsyLuban table template.")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--full-name", required=True, help="table full_name, e.g. test.TbExample")
    parser.add_argument("--value-type", help="override value_type; omit to derive it from the table name (TbExample -> Example)")
    parser.add_argument("--index", help="override index field; omit to use the first column")
    parser.add_argument("--mode", help="override mode (map/one/list); omit for map")
    parser.add_argument("--field", action="append", default=[], help="field spec name:type (repeatable)")
    args = parser.parse_args()

    output_path = Path(args.output)
    if output_path.suffix.lower() != ".xlsx":
        raise SystemExit("output must be .xlsx")

    fields = []
    for spec in args.field:
        if ":" not in spec:
            raise SystemExit(f"invalid --field '{spec}', expected name:type")
        name, ftype = spec.split(":", 1)
        fields.append((name.strip(), ftype.strip()))

    if not fields:
        fields = [(args.index or "id", "int"), ("name", "string")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # B1 carries only what cannot be inferred. full_name is the table's identity;
    # read_schema_from_file="true" is required because this template describes its
    # structure in the ##var / ##type rows rather than in a schema XML.
    # value_type / index / mode all have defaults -- writing them out unchanged
    # would be noise that every future reader has to check against the defaults.
    parts = [f'full_name="{args.full_name}"', 'read_schema_from_file="true"']
    if args.value_type:
        parts.insert(1, f'value_type="{args.value_type}"')
    if args.index:
        parts.append(f'index="{args.index}"')
    if args.mode:
        parts.append(f'mode="{args.mode}"')

    ws.cell(1, 1).value = "##export"
    ws.cell(1, 2).value = " & ".join(parts)

    ws.cell(2, 1).value = "##var"
    ws.cell(3, 1).value = "##type"

    for idx, (name, ftype) in enumerate(fields, start=2):
        ws.cell(2, idx).value = name
        ws.cell(3, idx).value = ftype

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"template created: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
